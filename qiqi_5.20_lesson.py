import openpyxl      #导入
import requests

#读取测试用例的函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename) #加载了这个工作簿---Excel表格==赋值给一个变量
    sheet = wb[sheetname]     #表单
    max_row = sheet.max_row   #获取最大行号
    cases = []         #空列表
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,  #获取序号
        url = sheet.cell(row=i,column=5).value,     #获取url
        data = sheet.cell(row=i,column=6).value,     #获取data
        expected_result = sheet.cell(row=i, column=7).value     #获取期望结果
        )          #一个用例放到一个字典
        cases.append(case)    #把字典追加到列表里保存
    return cases              #定义返回值
#函数---发送接口请求
def post_func(qcd_url,qcd_data):
    res = requests.post(url=qcd_url,data=qcd_data)        #post方法发送接口请求
    result = res.json()       #变量的申明
    return result           #返回值
#写入结果的方法
def write_result(filename,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(filename) #加载了这个工作簿---Excel表格==赋值给一个变量
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = real_result    #写入
    wb.save(filename)                #保存文件

def do_func(filename,sheetname):
    tese_cases = read_data(filename,sheetname)
    # print(tese_cases)
    for test_case in tese_cases:
        case_id = test_case.get('case_id')     #获取对应case_id
        url = test_case.get('url')            #获取对应url
        data = test_case.get('data')    #获取对应接口请求参数   ==  文本---字符串 ==  转化为字典格式
        data = eval(data)           #  eval()进行数据类型转化 --- 字符串->字典
        # print(data)          # data---必须是字典格式
        expected_result = test_case['expected_result']     #获取对应的期望结果
        expected_result = expected_result.replace('null','None')         #字符串的替换
        expected_result = eval(expected_result)           #  eval()进行数据类型转化 --- 字符串->字典
        #print(type(expected_result))
        real_result = post_func(qcd_url=url,qcd_data=data)         #调用接口发送函数---接口请求
        # print(type(real_result))
        real_msg = real_result.get('msg')       #字典取值---获取要断言的有效字段
        # print(real_msg)
        expected_msg = expected_result.get('msg')
        print('实际执行结果是:{}'.format(real_msg))
        print('预期测试结果是:{}'.format(expected_msg))
        if real_msg == expected_msg:
            print('第{}条测试用例通过'.format(case_id))
            final_result = 'passed'             #变量---目的：回写结果
        else:
            print('第{}条测试用例不通过'.format(case_id))
            final_result = 'failed'
        print('**'*20)
        write_result(filename,sheetname,case_id+1,8,final_result)        #调用了回写的函数--结果写入
do_func('test_case.xlsx','register')